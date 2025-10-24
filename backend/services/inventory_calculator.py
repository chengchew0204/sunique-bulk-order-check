import pandas as pd
import numpy as np
import requests
import msal
import time
from datetime import datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from concurrent.futures import ThreadPoolExecutor, as_completed


class InventoryCalculator:
    def __init__(self, config):
        self.config = config
        self.access_token = None
        self.site_id = None
        self.drive_id_documents = None
        self.folder_id = None
        
    def authenticate_sharepoint(self):
        """Authenticate with SharePoint and get access token"""
        authority = f"https://login.microsoftonline.com/{self.config['SHAREPOINT_TENANT_ID']}"
        scope = ['https://graph.microsoft.com/.default']
        
        app = msal.ConfidentialClientApplication(
            self.config['SHAREPOINT_CLIENT_ID'],
            authority=authority,
            client_credential=self.config['SHAREPOINT_CLIENT_SECRET']
        )
        
        result = app.acquire_token_for_client(scopes=scope)
        
        if "access_token" in result:
            self.access_token = result["access_token"]
            return True
        else:
            raise Exception(f"Failed to acquire token: {result.get('error_description')}")
    
    def get_site_info(self):
        """Get SharePoint site ID"""
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            'Content-Type': 'application/json'
        }
        
        site_url = f"https://graph.microsoft.com/v1.0/sites/{self.config['SHAREPOINT_HOSTNAME']}:/sites/{self.config['SHAREPOINT_SITE_NAME']}"
        response = requests.get(site_url, headers=headers)
        
        if response.status_code == 200:
            site_info = response.json()
            self.site_id = site_info['id']
            return True
        else:
            raise Exception(f"Failed to get Site ID: {response.status_code} {response.text}")
    
    def get_drive_id(self):
        """Get Documents library drive ID"""
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            'Content-Type': 'application/json'
        }
        
        drive_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives"
        response = requests.get(drive_url, headers=headers)
        
        if response.status_code == 200:
            drives = response.json()
            for drive in drives['value']:
                if drive['name'] == 'Documents':
                    self.drive_id_documents = drive['id']
                    return True
            raise Exception("Documents library not found")
        else:
            raise Exception(f"Failed to get Drive ID: {response.status_code} {response.text}")
    
    def get_folder_id(self):
        """Get Technology Stuff folder ID"""
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            'Content-Type': 'application/json'
        }
        
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{self.drive_id_documents}/root/children"
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            files = response.json()
            for item in files['value']:
                if item['name'] == 'Technology Stuff (DO NOT CHANGE)':
                    self.folder_id = item['id']
                    return True
            raise Exception("Technology Stuff folder not found")
        else:
            raise Exception(f"Failed to get folder list: {response.status_code} {response.text}")
    
    def _download_file(self, file_path, headers):
        """Download a single file from SharePoint"""
        download_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{self.drive_id_documents}/items/{file_path}/content"
        response = requests.get(download_url, headers=headers)
        if response.status_code != 200:
            raise Exception(f"Failed to download {file_path}: {response.status_code}")
        return response.content
    
    def download_sharepoint_data(self):
        """Download all required Excel files from SharePoint in parallel"""
        method_start = time.time()
        
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            'Content-Type': 'application/json'
        }
        
        # Get folder contents to find file IDs
        folder_scan_start = time.time()
        url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{self.drive_id_documents}/items/{self.folder_id}/children"
        response = requests.get(url, headers=headers)
        
        sales_order_checklist_id = None
        raw_data_folder_id = None
        
        if response.status_code == 200:
            items = response.json()
            for item in items['value']:
                if item['name'] == 'Sales Order Process Checklist.xlsx':
                    sales_order_checklist_id = item['id']
                elif item['name'] == 'Raw Data (Do Not Open in 7-8pm)':
                    raw_data_folder_id = item['id']
        else:
            raise Exception(f"Failed to get folder contents: {response.status_code} {response.text}")
        
        print(f"[PERF]   - Folder scan: {time.time() - folder_scan_start:.2f}s")
        
        # Create download tasks for all 4 files
        download_tasks = [
            ('sales_checklist', sales_order_checklist_id),
            ('inventory', f"{raw_data_folder_id}:/Inventory.xlsx:"),
            ('order', f"{raw_data_folder_id}:/Order.xlsx:"),
            ('ls_tracking', f"{self.folder_id}:/LS order tracking.xlsx:")
        ]
        
        # Download all files in parallel
        download_start = time.time()
        downloaded_data = {}
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {executor.submit(self._download_file, task[1], headers): task[0] 
                       for task in download_tasks}
            
            for future in as_completed(futures):
                name = futures[future]
                try:
                    downloaded_data[name] = future.result()
                except Exception as e:
                    raise Exception(f"Failed to download {name}: {str(e)}")
        
        print(f"[PERF]   - Parallel downloads: {time.time() - download_start:.2f}s")
        
        # Parse all Excel files from downloaded data
        parse_start = time.time()
        df_sales_order_check_list = pd.read_excel(BytesIO(downloaded_data['sales_checklist']))
        df_inventory = pd.read_excel(BytesIO(downloaded_data['inventory']))
        df_order = pd.read_excel(BytesIO(downloaded_data['order']))
        df_order_tracking = pd.read_excel(BytesIO(downloaded_data['ls_tracking']), sheet_name='Shipped')
        print(f"[PERF]   - Parse Excel files: {time.time() - parse_start:.2f}s")
        
        print(f"[PERF]   - download_sharepoint_data total: {time.time() - method_start:.2f}s")
        
        return {
            'sales_order_check_list': df_sales_order_check_list,
            'inventory': df_inventory,
            'order': df_order,
            'order_tracking': df_order_tracking
        }
    
    def process_order_tracking(self, df_order_tracking):
        """Extract stock on sea data"""
        df_order_tracking = df_order_tracking.iloc[3:, 1:]
        df_order_tracking = df_order_tracking.fillna(0)
        df_order_tracking.columns = df_order_tracking.iloc[0].tolist()
        df_order_tracking = df_order_tracking.drop(df_order_tracking.index[0]).reset_index(drop=True)
        df_order_tracking = df_order_tracking.rename(columns={'ProductName': 'SKU'})
        df_order_tracking["Total Stock On Sea"] = df_order_tracking.iloc[:, 1:].fillna(0).sum(axis=1)
        df_order_tracking = df_order_tracking[['SKU', 'Total Stock On Sea']]
        return df_order_tracking
    
    def calculate_sales_forecast(self, df_order, df_sales_order_check_list, df_inventory):
        """Calculate 2-month sales forecast from 3-month history"""
        current_date = datetime.now()
        past_month = current_date - relativedelta(months=3)
        current_date_str = current_date.strftime('%Y-%m-%d')
        past_month_str = past_month.strftime('%Y-%m-%d')
        
        # Filter orders
        df_order = df_order[
            df_order["Order Date"].notna()
            & (df_order["Order Date"].astype(str).str.strip() != "")
            & (df_order["Order Status"] != "quote")
            & (df_order["Order Status"] != "unconfirmed")
        ]
        df_order["Order Date"] = pd.to_datetime(df_order["Order Date"], errors="coerce")
        df_order = df_order[(df_order["Order Date"] >= past_month_str) & (df_order["Order Date"] <= current_date_str)]
        df_order = df_order[['Order Number', 'Product Name', 'Product Quantity']].rename(columns={'Order Number': 'Order #'})
        
        # Process sales order checklist
        df_sales_order_check_list = df_sales_order_check_list[['Order #', 'Order Date', 'Bulk Order']]
        cols = ["Order Date", "Bulk Order"]
        df_sales_order_check_list[cols] = df_sales_order_check_list[cols].astype(str).apply(lambda x: x.str.lstrip("'"))
        df_sales_order_check_list = df_sales_order_check_list[
            df_sales_order_check_list["Order Date"].notna() & (df_sales_order_check_list["Bulk Order"] != "Yes")
        ]
        df_sales_order_check_list["Order Date"] = pd.to_datetime(df_sales_order_check_list["Order Date"], errors="coerce")
        df_sales_order_check_list = df_sales_order_check_list[
            (df_sales_order_check_list["Order Date"] >= past_month_str) & 
            (df_sales_order_check_list["Order Date"] <= current_date_str)
        ]
        
        # Merge and filter
        df_sales_order_check_list = pd.merge(df_sales_order_check_list, df_order, how="left", on="Order #")
        pattern = r'FG-|GB-|SSW-|SW-|ET-|SSO-|SWB-'
        df_sales_order_check_list = df_sales_order_check_list[
            df_sales_order_check_list["Product Name"].str.contains(pattern, regex=True, na=False)
        ]
        df_sales_order_check_list = df_sales_order_check_list.groupby("Product Name", as_index=False)["Product Quantity"].sum()
        
        # Merge with inventory
        df_inventory['rawQuantityAvailable'] = pd.to_numeric(df_inventory['rawQuantityAvailable'], errors='coerce')
        df_inventory_temp = df_inventory[['name', 'rawQuantityAvailable']].rename(columns={'name': 'Product Name'})
        df_sales_order = pd.merge(df_inventory_temp, df_sales_order_check_list, on='Product Name', how='left')
        df_sales_order = df_sales_order.rename(columns={'Product Quantity': 'sale_amount_in_two_months'})
        df_sales_order = df_sales_order[~(df_sales_order['sale_amount_in_two_months'] == 0)]
        df_sales_order = df_sales_order[df_sales_order["sale_amount_in_two_months"].notna()]
        df_sales_order['sale_amount_in_two_months'] = np.ceil(df_sales_order['sale_amount_in_two_months'] / 3 * 2).astype(int)
        
        return df_inventory_temp, df_sales_order
    
    def update_stock(self, df_main, df_other):
        """Add in-transit stock for items with 0-4 months inventory"""
        mask = (df_main["result_temp"] > 0) & (df_main["result_temp"] < 4)
        
        for idx, row in df_main[mask].iterrows():
            sku = row["SKU"]
            match = df_other[df_other["SKU"] == sku]
            
            if not match.empty:
                stock_other = match["Total Stock On Sea"].iloc[0]
                df_main.at[idx, "Stock"] = row["Stock"] + stock_other
        
        return df_main
    
    def calc_need(self, row):
        """Calculate adjusted NEED to maintain 4-month safety stock"""
        need = row['NEED']
        res = row['result']
        stk = row['Stock']
        sale = row['Sale']
        
        if res >= 4:
            return need
        elif res > 0:
            temp = int(stk - sale * 4)
            if (temp % 5 != 0):
                val = temp - (temp % 5)
            else:
                val = temp
            return max(val, 0)
        else:
            return 0
    
    def generate_results(self, df_bulk_order):
        """Main orchestration function"""
        start_time = time.time()
        print(f"[PERF] Starting inventory calculation...")
        
        # Authenticate and download data
        auth_start = time.time()
        self.authenticate_sharepoint()
        print(f"[PERF] Authentication: {time.time() - auth_start:.2f}s")
        
        site_start = time.time()
        self.get_site_info()
        print(f"[PERF] Get site info: {time.time() - site_start:.2f}s")
        
        drive_start = time.time()
        self.get_drive_id()
        print(f"[PERF] Get drive ID: {time.time() - drive_start:.2f}s")
        
        folder_start = time.time()
        self.get_folder_id()
        print(f"[PERF] Get folder ID: {time.time() - folder_start:.2f}s")
        
        download_start = time.time()
        data = self.download_sharepoint_data()
        print(f"[PERF] Download all files: {time.time() - download_start:.2f}s")
        
        # Process order tracking
        tracking_start = time.time()
        df_order_tracking = self.process_order_tracking(data['order_tracking'])
        print(f"[PERF] Process order tracking: {time.time() - tracking_start:.2f}s")
        
        # Calculate sales forecast
        forecast_start = time.time()
        df_inventory_temp, df_sales_order = self.calculate_sales_forecast(
            data['order'],
            data['sales_order_check_list'],
            data['inventory']
        )
        print(f"[PERF] Calculate sales forecast: {time.time() - forecast_start:.2f}s")
        
        # Extract stock and sales, merge, and calculate results
        calc_start = time.time()
        
        # Extract stock
        stock = df_inventory_temp.rename(columns={'Product Name': 'SKU', 'rawQuantityAvailable': 'Stock'})
        stock['Stock'] = pd.to_numeric(stock['Stock'], errors='coerce').fillna(0).astype(int)
        
        # Extract sales
        sale = df_sales_order[['Product Name', 'sale_amount_in_two_months']].rename(
            columns={'Product Name': 'SKU', 'sale_amount_in_two_months': 'Sale'}
        )
        sale['Sale'] = sale['Sale'].fillna(0)
        sale['Sale'] = pd.to_numeric(sale['Sale'], errors='coerce').clip(lower=0)
        sale['Sale'] = (sale['Sale'] / 2 * 1.5).astype(int)
        
        # Merge with bulk order
        df_result = df_bulk_order.merge(stock, on="SKU", how="left").merge(sale, on='SKU', how='left')
        df_result["Stock"] = df_result["Stock"].fillna(0).astype(int)
        df_result["Sale"] = df_result["Sale"].fillna(0).astype(int)
        
        # Calculate temporary result
        df_result["result_temp"] = (df_result["Stock"] - df_result["NEED"]) / df_result["Sale"]
        df_result["result_temp"] = df_result["result_temp"].replace([np.inf, -np.inf], 0)
        
        # Update stock with in-transit inventory
        df_result = self.update_stock(df_result, df_order_tracking)
        
        # Calculate final result
        df_result["result"] = (df_result["Stock"] - df_result["NEED"]) / df_result["Sale"]
        df_result["result"] = df_result["result"].replace([np.inf, -np.inf], 0)
        df_result.drop(columns=["result_temp"], inplace=True)
        
        # Calculate actual can sell
        df_result['Actual Can Sell'] = df_result.apply(self.calc_need, axis=1)
        
        # Select final columns in correct order: SKU, NEED, result, Stock, Sale, Actual Can Sell
        df_final = df_result[['SKU', 'NEED', 'result', 'Stock', 'Sale', 'Actual Can Sell']].copy()
        
        print(f"[PERF] Final calculations: {time.time() - calc_start:.2f}s")
        print(f"[PERF] TOTAL TIME: {time.time() - start_time:.2f}s")
        
        return df_final
    
    def generate_excel_with_formatting(self, df, output_path):
        """Generate Excel file with conditional formatting"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            wb = writer.book
            ws = writer.sheets['Sheet1']
            
            # Define fill styles
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            n = len(df)
            
            # Apply conditional formatting to SKU column (A) based on result column (C)
            # Green: result >= 4 (Safe - 4+ months inventory)
            ws.conditional_formatting.add(
                f'A2:A{n+1}',
                FormulaRule(formula=['$C2>=4'], fill=green_fill)
            )
            # Yellow: 0 < result < 4 (Warning - low inventory)
            ws.conditional_formatting.add(
                f'A2:A{n+1}',
                FormulaRule(formula=['AND($C2>0,$C2<4)'], fill=yellow_fill)
            )
            # Red: result <= 0 (Critical - out of stock)
            ws.conditional_formatting.add(
                f'A2:A{n+1}',
                FormulaRule(formula=['$C2<=0'], fill=red_fill)
            )

