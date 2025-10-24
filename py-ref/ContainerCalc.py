# -*- coding: utf-8 -*-
"""
Created on Tue Jul 22 14:01:13 2025

@author: 15253
"""

import pandas as pd
import numpy as np
import requests
from datetime import datetime
from dateutil.relativedelta import relativedelta
import time
import os
#import pickle
#from google.auth.transport.requests import Request
#from google_auth_oauthlib.flow import InstalledAppFlow
import msal
import base64
from openpyxl import load_workbook
from pandas.tseries.offsets import DateOffset
import json
import concurrent.futures
from io import BytesIO
from dateutil import parser
import traceback
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd



###Cabinet###
companyID_cabinet = '9a253611-78ee-445e-90ed-f6c9f0e7b6fe' #Current using account
API_cabinet = '41AB3DDAE9DE173187A5AA03DD93326D6C39B22AE02B8593E361F165AACCBFBD-1' #Current using account
###Cabinet###


'''
Backend的密钥和凭证
'''
###以下为Backend发送邮件的Client和Tenant信息##
client_secret = 'your_client_secret_here'
tenant_id = 'your_tenant_id_here'
client_secret_value = 'your_client_secret_value_here'
client_id = 'your_client_id_here'
application_client_id_email = 'your_application_client_id_here'
###以上为Backend发送邮件的Client和Tenant信息###

authority = f"https://login.microsoftonline.com/{tenant_id}"
scopes = ['https://graph.microsoft.com/.default']
 

###以下为与sharepoint交互的Client和Tenant信息###
client_value = 'your_client_value_here'#有效期720天 从3/30/2025开始
client_secret_sharepoint = 'your_client_secret_sharepoint_here'#有效期720天 从3/30/2025开始
tenant_id = 'your_tenant_id_here'
application_client_id = 'your_application_client_id_here'
object_id = 'your_object_id_here'
###以上为与sharepoint交互的Client和Tenant信息###

authority = f"https://login.microsoftonline.com/{tenant_id}"
scope = ['https://graph.microsoft.com/.default']

'''
与sharepoint中的excel文档交互
'''
#获取访问令牌
app = msal.ConfidentialClientApplication(application_client_id, authority=authority, client_credential=client_value)
result = app.acquire_token_for_client(scopes=scope)

if "access_token" in result:
    access_token = result["access_token"]
    #print("获取访问令牌成功")
else:
    print("获取访问令牌失败:", result.get("error_description"))

headers_sharepoint = {
    "Authorization": f"Bearer {access_token}",
    'Content-Type': 'application/json'
}  

hostname = 'suniquecabinetry.sharepoint.com'
site_name = 'sccr'

# 请求 Site ID
site_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/sites/{site_name}"
response = requests.get(site_url, headers=headers_sharepoint)

if response.status_code == 200:
    site_info = response.json()
    site_id = site_info['id']
    #print("Site ID:", site_id)
else:
    print("获取 Site ID 失败:", response.status_code, response.text)



# 通过 Site ID 获取文档库（Drive）ID
drive_id_documents = ''
drive_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
response = requests.get(drive_url, headers=headers_sharepoint)

if response.status_code == 200:
    drives = response.json()
    for drive in drives['value']:
        if drive['name'] == 'Documents':
            drive_id_documents = drive['id']
        #print("Drive ID:", drive['id'], "Drive Name:", drive['name'])
else:
    print("获取 Drive ID 失败:", response.status_code, response.text)

# 列出文档库根目录的文件或者文件夹，此处为获取Technology Stuff (DO NOT CHANGE)的ID和绝对路径
folder_path = ''
folder_id = ''
url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id_documents}/root/children"
response = requests.get(url, headers=headers_sharepoint)

if response.status_code == 200:
    files = response.json()
    for item in files['value']:
        if item['name'] == 'Technology Stuff (DO NOT CHANGE)':
            folder_id = item['id']
            folder_path = item['parentReference']['path']
        #print(f"Name: {item['name']}, File ID: {item['id']}, Path: {item['parentReference']['path']}")
else:
    print("获取文件列表失败:", response.status_code, response.text)

# 获取文件夹中的子文件和文件夹
url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id_documents}/items/{folder_id}/children"
response = requests.get(url, headers=headers_sharepoint)
#Sales Order Process Checklist
Sales_Order_Checklist_id = ''
Sales_Order_Checklist_path = ''
#Raw data folder
raw_data_folder_id = ''
raw_data_folder_path = ''
#Order Ready List
order_ready_list_id = ''
order_ready_list_path = ''

if response.status_code == 200:
    items = response.json()
    for item in items['value']:
        if item['name'] == 'Sales Order Process Checklist.xlsx':
            Sales_Order_Checklist_id = item['id']
            Sales_Order_Checklist_path = item['parentReference']['path']
        elif item['name'] == 'Raw Data (Do Not Open in 7-8pm)':
            raw_data_folder_id = item['id']
            raw_data_folder_path = item['parentReference']['path']
        elif item['name'] == 'Order Ready List.xlsx':
            order_ready_list_id = item['id']
            order_ready_list_path = item['parentReference']['path']
        #print(f"Name: {item['name']}, ID: {item['id']}, Type: {item['folder'] if 'folder' in item else 'file'}")
else:
    print("获取文件夹内容失败:", response.status_code, response.text)
    
#下载最新的sales order list
download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id_documents}/items/{Sales_Order_Checklist_id}/content"
#local_file_path_checklist = "C:/Users/paopa/OneDrive/Desktop/Raw Data/Sales Order Process Checklist.xlsx"
#file_name = "Sales Order Process Checklist.xlsx"  # Name of the file on SharePoint
response = requests.get(download_url, headers=headers_sharepoint)
if response.status_code == 200:
    sales_order_check_list = BytesIO(response.content)
    df_sales_order_check_list = pd.read_excel(sales_order_check_list)
else:
    print("Failed to download sales order checklist:", response.status_code, response.text)
    
#下载最新的Inventroy list    
download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id_documents}/items/{raw_data_folder_id}:/Inventory.xlsx:/content"

response = requests.get(download_url, headers=headers_sharepoint)
if response.status_code == 200:
    inventroy = BytesIO(response.content)
    df_inventroy = pd.read_excel(inventroy)
else:
    print("Failed to download sales order checklist:", response.status_code, response.text)
    
#下载最新的Order list    
download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id_documents}/items/{raw_data_folder_id}:/Order.xlsx:/content"

response = requests.get(download_url, headers=headers_sharepoint)
if response.status_code == 200:
    Order = BytesIO(response.content)
    df_order = pd.read_excel(Order)
else:
    print("Failed to download sales order checklist:", response.status_code, response.text)

download_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives/{drive_id_documents}/items/{folder_id}:/LS order tracking.xlsx:/content"
headers = {
    "Authorization": f"Bearer {access_token}",
    'Content-Type': 'application/json'
}
# local_file_path_checklist = r"C:/Users/paopa/OneDrive/Desktop/python\LS order tracking.xlsx"
file_name = "LS order tracking.xlsx"  # Name of the file on SharePoint
response = requests.get(download_url, headers=headers)
if response.status_code == 200:
    LS_order_tracking = BytesIO(response.content)
    df_order_tracking = pd.read_excel(LS_order_tracking,sheet_name='Shipped')
    #print(f"Template file downloaded successfully and saved as {local_file_path_checklist}")
else:
    print("Failed to download the template file:", response.status_code, response.text)
    
#去除多余行
df_order_tracking = df_order_tracking.iloc[3:, 1:]
#去除含NAN的列
df_order_tracking = df_order_tracking.fillna(0)
df_order_tracking.columns = df_order_tracking.iloc[0].tolist()
#将第一列的名字给index赋值
df_order_tracking = df_order_tracking.drop(df_order_tracking.index[0]).reset_index(drop=True)
df_order_tracking = df_order_tracking.rename(columns={'ProductName': 'SKU'})
df_order_tracking["Total Stock On Sea"] = df_order_tracking.iloc[:, 1:].fillna(0).sum(axis=1)
df_order_tracking = df_order_tracking[['SKU','Total Stock On Sea']]

current_date = datetime.now()
past_month = current_date - relativedelta(months=3)
current_date = current_date.strftime('%Y-%m-%d')
past_month = past_month.strftime('%Y-%m-%d')
# base_url = f'https://cloudapi.inflowinventory.com/{companyID_cabinet}/sales-orders?count=100&include=lines&filter[orderDate]={{"fromDate":"{past_month}", "toDate":"{current_date}"}}'


def update_stock(df_main, df_other):
    # 只处理 0 < result < 4 的行
    mask = (df_main["result_temp"] > 0) & (df_main["result_temp"] < 4)

    for idx, row in df_main[mask].iterrows():
        sku = row["SKU"]

        # 在 df_other 中找到相同 SKU
        match = df_other[df_other["SKU"] == sku]

        if not match.empty:
            stock_other = match["Total Stock On Sea"].iloc[0]  # 取 df_other 中的 stock
            df_main.at[idx, "Stock"] = row["Stock"] + stock_other  # 相加并更新

    return df_main


# df = pd.DataFrame()
# total = 99999  #如果总数据有913个 设置1100能捕获全部
# limit = 100 #API上限 一次最多捕获100个数据
# counter = 1

# #获取1个月内的订单
# """Cabinet low stock"""
# headers = { 'Authorization': 'Bearer ' + API_cabinet,
#             'Content-Type': 'application/json',
#             'Accept': 'application/json;version=2024-03-12',
#             'X-OverrideAllowNegativeInventory': 'TRUE'
#           }

# while counter < (total // 100):
    
#     if counter == 1:
#         response = requests.get(base_url,headers = headers)
#         if response.status_code != 200:       
#             break
    
#         json_assets = response.json()
#         productList = pd.json_normalize(json_assets)
#         df = pd.concat([df,productList],ignore_index=True)
        
#         counter +=1
        
#     else:
#         response = requests.get(f'https://cloudapi.inflowinventory.com/{companyID_cabinet}/sales-orders?count=100&include=lines&filter[orderDate]={{"fromDate":"{past_month}", "toDate":"{current_date}"}}&skip={limit}', headers = headers)
#         if response.status_code != 200:
#             break
        
#         json_assets = response.json()
#         productList = pd.json_normalize(json_assets)
#         if productList.empty:
#             break
#         df = pd.concat([df,productList],ignore_index=True)
         
#         counter +=1
#         limit += 100


# #筛选出Invoiced，Owing，Paid和Partial四种支付状态的订单
# filtered_df_payment = df[df['paymentStatus'].isin(['invoiced', 'owing', 'paid', 'partial'])]
# #删除为bulk order的订单
# filtered_df_payment = filtered_df_payment[~filtered_df_payment['customFields.custom5'].str.contains(r'[a-zA-Z]', na=False)]

# df_exploded = filtered_df_payment.explode('lines')
# #单独抓出包含订单产品信息的栏
# df_expanded = pd.json_normalize(df_exploded['lines'])
# df_expanded['quantity'] = pd.to_numeric(df_expanded['quantity.standardQuantity'], errors='coerce')
# #单独抓出产品ID和下单数量
# df_grouped = df_expanded.groupby('productId', as_index=False).agg({
#     'quantity': 'sum',
# })

# #根据订单中的product ID 获取产品的信息 寻找是否包含FG|GB|SSW|SW
# df_new = pd.DataFrame()
# temp_counter = 0
# for product_id in df_grouped['productId']:
#     print(temp_counter)
#     temp_counter += 1
#     response = requests.get(f'https://cloudapi.inflowinventory.com/{companyID_cabinet}/products/{product_id}?include=category', headers = headers)
#     if response.status_code != 200:
#         while response.status_code == 429:
#             time.sleep(40)
#             response = requests.get(f'https://cloudapi.inflowinventory.com/{companyID_cabinet}/products/{product_id}?include=category', headers = headers)
#             if(response.status_code == 200):
#                 break

#     json_assets = response.json()
#     productList_2 = pd.json_normalize(json_assets)
#     df_new = pd.concat([df_new,productList_2],ignore_index=True)



# #筛选出所需产品的SKU
# #随着库存的更新而更新
# #FG|GB|SSW|SW|SWB 10/15/2024


# df_new['name'] = df_new['name'].astype(str)
# df_new['name'] = df_new['name'].str.strip()


# #检查是否为stockedProduct
# stocked_product_id = []
# for index, product_id in enumerate(df_new['productId']):
#     if(df_new['itemType'][index] == 'stockedProduct'):
#         stocked_product_id.append(product_id)
    


# df2 = pd.DataFrame()
# #根据筛选后的产品ID 获取产品的详细信息
# for product_id in stocked_product_id:
#     response = requests.get(f'https://cloudapi.inflowinventory.com/{companyID_cabinet}/products/{product_id}/summary', headers = headers,timeout=60)
#     if response.status_code != 200:
#         while response.status_code == 429:
#             print(product_id)
#             time.sleep(45)
#             response = requests.get(f'https://cloudapi.inflowinventory.com/{companyID_cabinet}/products/{product_id}/summary', headers = headers,timeout=60)
#             if(response.status_code == 200):
#                 break
#         else:
#             print('Error' + str(response.status_code))
    
#     json_assets = response.json()
#     productSummary = pd.json_normalize(json_assets)
#     df2 = pd.concat([df2,productSummary],ignore_index=True)
    
# #预估2个月内的销量 查询现有库存减去两个月销量是否大于0 若不够则加入补货列表
# df2['quantityAvailable'] = pd.to_numeric(df2['quantityAvailable'], errors='coerce')
# df2 = df2[['productId','quantityAvailable']]
# #按照productID将库存和预计销量合并
# df2 = pd.merge(df2, df_grouped, on='productId', how='left')
# df2 = df2.rename(columns={'quantity': 'sale_amount_in_two_months'})
# '''以下为3/25改过的部分'''
# df2 = df2[~(df2['sale_amount_in_two_months'] == 0)]
# '''以上为3/25改过的部分'''
# df2['sale_amount_in_two_months'] = np.ceil(df2['sale_amount_in_two_months'] / 3 * 2).astype(int)
# #整理Data Frame 只留下需要补货的产品名称，现可用数量和预计2个月后的库存数量
# merged_df = pd.merge(df2, df_new[['productId', 'name']], on='productId', how='left')  
# result_df = merged_df[['name', 'quantityAvailable','sale_amount_in_two_months']]
# result_df.to_excel(fr"C:\Users\15253\Desktop\python\Container\Average Sale {current_date}.xlsx", index=False)


# 读取三张 Excel
df_test_sale = pd.read_excel(r"/Users/zackwu204/CursorAI/Sunique/03-bulk-order-ckecking/development-local/data/BulkOrder.xlsx")
# df_average_sale = pd.read_excel(r"C:\Users\15253\Desktop\python\Container\Average Sale.xlsx",sheet_name='Sheet1')
# df_stock = pd.read_excel(r"C:\Users\15253\Desktop\python\Container\Stock.xlsx",sheet_name='Stock')
#df_price = pd.read_excel(r"C:\Users\15253\Desktop\工作用的\Container\ARK Profoma invoice PO#ARK02（14.06.2025）.xlsx")

"""Cabinet low stock"""
df_order = df_order[ df_order["Order Date"].notna()
                    & (df_order["Order Date"].astype(str).str.strip() != "")
                    & (df_order["Order Status"] != "quote")
                    & (df_order["Order Status"] != "unconfirmed")]
df_order["Order Date"] = pd.to_datetime(df_order["Order Date"], errors="coerce")
df_order = df_order[(df_order["Order Date"] >= past_month) & (df_order["Order Date"] <= current_date)]
df_order = df_order[['Order Number','Product Name','Product Quantity']].rename(columns={
                                'Order Number': 'Order #'
                                })    
    
df_sales_order_check_list = df_sales_order_check_list[['Order #','Order Date','Bulk Order']]
cols = ["Order Date", "Bulk Order"]
df_sales_order_check_list[cols] = df_sales_order_check_list[cols].astype(str).apply(lambda x: x.str.lstrip("'"))
df_sales_order_check_list = df_sales_order_check_list[ df_sales_order_check_list["Order Date"].notna() & (df_sales_order_check_list["Bulk Order"] != "Yes") ]
# 确保 df 的 orderDate 是 datetime 类型
df_sales_order_check_list["Order Date"] = pd.to_datetime(df_sales_order_check_list["Order Date"], errors="coerce")
# 直接用 datetime 对象做筛选
df_sales_order_check_list = df_sales_order_check_list[(df_sales_order_check_list["Order Date"] >= past_month) & (df_sales_order_check_list["Order Date"] <= current_date)]

df_sales_order_check_list = pd.merge(df_sales_order_check_list, df_order, how="left", left_on="Order #", right_on="Order #")

pattern = r'FG-|GB-|SSW-|SW-|ET-|SSO-|SWB-'
df_sales_order_check_list = df_sales_order_check_list[df_sales_order_check_list["Product Name"].str.contains(pattern, regex=True, na=False)]
df_sales_order_check_list = df_sales_order_check_list.groupby("Product Name", as_index=False)["Product Quantity"].sum()

df_inventroy['rawQuantityAvailable'] = pd.to_numeric(df_inventroy['rawQuantityAvailable'], errors='coerce')
df_inventroy_temp = df_inventroy[['name','rawQuantityAvailable']].rename(columns={
                                'name': 'Product Name'
                                })    
#按照productID将库存和预计销量合并
df_sales_order = pd.merge(df_inventroy_temp, df_sales_order_check_list, on='Product Name', how='left')
df_sales_order = df_sales_order.rename(columns={'Product Quantity': 'sale_amount_in_two_months'})
'''以下为3/25改过的部分'''
df_sales_order = df_sales_order[~(df_sales_order['sale_amount_in_two_months'] == 0)]
df_sales_order = df_sales_order[ df_sales_order["sale_amount_in_two_months"].notna()]
'''以上为3/25改过的部分'''
df_sales_order['sale_amount_in_two_months'] = np.ceil(df_sales_order['sale_amount_in_two_months'] / 3 * 2).astype(int)

# #在主表里提取标准SKU，比如把 "[SW-3DB21] SW-3DB21" → "SW-3DB21"
# df_test_sale["SKU"] = (
#     df_test_sale["SKU"]
#       .str.replace(r"[\[\]]", "", regex=True)   # 去掉中括号
#       .str.split()                              # 拆成 ["SW-3DB21", "SW-3DB21"]
#       .str[0]                                   # 取第一个
# )

# 3️⃣ 定义一个函数，用来从表2/表3的第一列提取 key 和尾部数字
def extract_stock(df):
    sku_col = df.columns[0]   # 第一列列名，存 SKU
    num_col = df.columns[1]   # 第二列列名，存数量

    tmp = df[[sku_col, num_col]].copy()
    # 直接把 SKU 作为 key（如果 SKU 上还带其他字符，需要再 .str.extract）
    tmp["name"] = tmp[sku_col].astype(str)
    # 数量强制转数字，不能转的当 0
    tmp["rawQuantityAvailable"] = pd.to_numeric(tmp[num_col], errors="coerce").fillna(0).astype(int)
    return tmp[["name", "rawQuantityAvailable"]]

# 假设 df2/df3 的第一列列名都是一样的，比如 "item"
stock = extract_stock(df_inventroy_temp)
stock = stock.rename(columns={
                    "name":'SKU',
                    "rawQuantityAvailable": "Stock"})

sale = df_sales_order[['Product Name','sale_amount_in_two_months']]
sale = df_sales_order.rename(columns={'Product Name':'SKU',
                                        "sale_amount_in_two_months": "Sale"})
sale['Sale'] = sale['Sale'].fillna(0)
sale['Sale'] = pd.to_numeric(sale['Sale'], errors='coerce').clip(lower=0)
sale['Sale'] = (sale['Sale'] / 2 * 1.5).astype(int)


df_test_sale = (
    df_test_sale
      .merge(stock, on="SKU", how="left")
      .merge(sale, on='SKU', how='left')
)

df_test_sale["Stock"] = df_test_sale["Stock"].fillna(0).astype(int)
df_test_sale["Sale"] = df_test_sale["Sale"].fillna(0).astype(int)
'''临时Result 用于计算是否需要添加在运输的橱柜到Stock中'''
df_test_sale["result_temp"] = (df_test_sale["Stock"] - df_test_sale["NEED"]) / (df_test_sale["Sale"])

df_test_sale = update_stock(df_test_sale, df_order_tracking)

df_test_sale["result"] = (df_test_sale["Stock"] - df_test_sale["NEED"]) / (df_test_sale["Sale"])
df_test_sale.drop(columns=["result_temp"], inplace=True)

# 根据 result 计算新的 NEED
def calc_need(row):
    
    need = row['NEED']
    res  = row['result']
    stk  = row['Stock']
    sale = row['Sale']
    
    if res >= 4:
        return need
    elif res > 0:
        temp = int(stk - sale * 4)
        if (temp % 5 != 0):
            val = temp - (temp % 5)
        # 防止负数
        else:
            val = temp
        return max(val,0)
    else:
        return 0
    

    
df_sku_count = df_test_sale[['SKU','NEED','result','Stock','Sale']].copy()
df_sku_count['Actual Can Sell'] = df_sku_count.apply(calc_need, axis=1)


with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    df_sku_count.to_excel(writer, index=False, sheet_name='Sheet1')
    wb = writer.book
    ws = writer.sheets['Sheet1']

    # 定义填充样式
    green_fill  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    n = len(df_sku_count)

    # 公式条件格式：基于 C 列（result）判断，给 A 列（SKU）上色
    ws.conditional_formatting.add(
        f'A2:A{n+1}',
        FormulaRule(formula=['$C2>=4'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f'A2:A{n+1}',
        FormulaRule(formula=['AND($C2>0,$C2<4)'], fill=yellow_fill)
    )
    ws.conditional_formatting.add(
        f'A2:A{n+1}',
        FormulaRule(formula=['$C2<=0'], fill=red_fill)
    )